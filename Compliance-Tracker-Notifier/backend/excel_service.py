import openpyxl
from datetime import datetime
from typing import List, Dict, Any, Optional
from sqlalchemy.orm import Session
from models import Project, ControlTemplate, ProjectControl, TaskInstance
import re


class ExcelImportService:
    """Service to import SPL 2.1 Excel files"""

    def __init__(self, db: Optional[Session]):
        self.db = db

    def get_sheet_names(self, file_path: str) -> List[str]:
        """Return all sheet names from the uploaded workbook"""
        workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        return workbook.sheetnames

    def parse_excel(self, file_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """Parse Excel file and extract control data from the selected sheet"""
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' was not found in the workbook")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        header_row_index, headers = self._detect_headers(sheet)
        controls = []
        current_control = None

        for row_number, row in enumerate(
            sheet.iter_rows(min_row=header_row_index + 1, values_only=True),
            start=header_row_index + 1
        ):
            if not row or not any(cell not in (None, '') for cell in row):
                continue

            control_objective = self._get_cell_value(row, headers, 'control_objective')
            control_execution_tasks = self._get_cell_value(row, headers, 'control_execution_tasks')
            control_guidance = self._get_cell_value(row, headers, 'control_guidance')
            frequency_event = self._get_cell_value(row, headers, 'frequency_event_driven')
            frequency_scheduled = self._get_cell_value(row, headers, 'frequency_scheduled')
            planned_date = self._get_cell_value(row, headers, 'planned_completion_date')
            actual_date = self._get_cell_value(row, headers, 'actual_completion_date')
            assigned_to = self._get_cell_value(row, headers, 'assigned_to')
            evidence_location = self._get_cell_value(row, headers, 'evidence_storage_location')

            normalized_objective = self._clean_text(control_objective)
            normalized_task = self._clean_text(control_execution_tasks)
            normalized_guidance = self._clean_text(control_guidance)
            normalized_frequency_event = self._clean_text(frequency_event)
            normalized_frequency_scheduled = self._clean_text(frequency_scheduled)
            normalized_evidence_location = self._clean_text(evidence_location)

            if normalized_objective:
                if current_control:
                    controls.append(current_control)

                current_control = {
                    'title': normalized_objective,
                    'description': normalized_task or '',
                    'control_code': normalized_guidance or '',
                    'frequency_event': normalized_frequency_event,
                    'frequency_scheduled': normalized_frequency_scheduled,
                    'assigned_to': self._parse_team_members(assigned_to),
                    'evidence_location': normalized_evidence_location or '',
                    'source_sheet': sheet.title,
                    'source_row_number': row_number,
                    'raw_task_name': self._stringify_cell(control_objective),
                    'raw_task_description': self._stringify_cell(control_execution_tasks),
                    'raw_guidance': self._stringify_cell(control_guidance),
                    'raw_frequency_event': self._stringify_cell(frequency_event),
                    'raw_frequency_scheduled': self._stringify_cell(frequency_scheduled),
                    'raw_assigned_to': self._stringify_cell(assigned_to),
                    'raw_evidence_location': self._stringify_cell(evidence_location),
                    'instances': []
                }
            elif current_control:
                if normalized_task and not current_control['description']:
                    current_control['description'] = normalized_task
                if normalized_guidance and not current_control['control_code']:
                    current_control['control_code'] = normalized_guidance
                if normalized_frequency_event and not current_control['frequency_event']:
                    current_control['frequency_event'] = normalized_frequency_event
                if normalized_frequency_scheduled and not current_control['frequency_scheduled']:
                    current_control['frequency_scheduled'] = normalized_frequency_scheduled
                if normalized_evidence_location and not current_control['evidence_location']:
                    current_control['evidence_location'] = normalized_evidence_location

                parsed_members = self._parse_team_members(assigned_to)
                if parsed_members and not current_control['assigned_to']:
                    current_control['assigned_to'] = parsed_members

            planned_date_text = self._stringify_cell(planned_date)
            actual_date_text = self._stringify_cell(actual_date)
            is_not_applicable = planned_date_text.strip().lower() in {'na', 'n/a'}

            parsed_planned_date = self._parse_date(planned_date)
            parsed_actual_date = self._parse_date(actual_date) if actual_date else None

            if current_control and (parsed_planned_date or is_not_applicable):
                effective_planned_date = parsed_planned_date or datetime.utcnow()
                effective_status = 'not-applicable' if is_not_applicable else ('completed' if parsed_actual_date else 'pending')

                instance = {
                    'planned_date': effective_planned_date,
                    'actual_date': parsed_actual_date,
                    'instance_label': self._generate_instance_label(
                        effective_planned_date,
                        current_control.get('frequency_scheduled')
                    ),
                    'status': effective_status,
                    'source_sheet': current_control.get('source_sheet'),
                    'source_row_number': row_number,
                    'raw_task_name': self._stringify_cell(control_objective),
                    'raw_task_description': self._stringify_cell(control_execution_tasks),
                    'raw_guidance': self._stringify_cell(control_guidance),
                    'raw_frequency_event': self._stringify_cell(frequency_event),
                    'raw_frequency_scheduled': self._stringify_cell(frequency_scheduled),
                    'raw_assigned_to': self._stringify_cell(assigned_to),
                    'raw_evidence_location': self._stringify_cell(evidence_location),
                    'raw_planned_date': planned_date_text,
                    'raw_actual_date': actual_date_text
                }
                current_control['instances'].append(instance)

        if current_control:
            controls.append(current_control)

        controls = [control for control in controls if control['title'] or control['instances']]

        return {
            'controls': controls,
            'total_controls': len(controls),
            'total_instances': sum(len(c['instances']) for c in controls)
        }

    def import_to_project(self, project_id: int, file_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """Import Excel data to a specific project"""
        if self.db is None:
            raise ValueError("Database session is required for importing data")

        try:
            parsed_data = self.parse_excel(file_path, sheet_name=sheet_name)

            project = self.db.query(Project).filter(Project.id == project_id).first()
            if not project:
                return {
                    'success': False,
                    'message': 'Project not found',
                    'controls_created': 0,
                    'tasks_created': 0,
                    'errors': ['Project not found']
                }

            controls_created = 0
            tasks_created = 0
            errors = []

            for control_data in parsed_data['controls']:
                try:
                    control_template = self._get_or_create_control_template(control_data)

                    project_control = self._create_project_control(
                        project_id,
                        control_template.id,
                        control_data
                    )

                    controls_created += 1

                    for instance_data in control_data['instances']:
                        self._create_task_instance(
                            project_control.id,
                            project_id,
                            control_template.id,
                            instance_data
                        )
                        tasks_created += 1

                except Exception as e:
                    errors.append(f"Error importing control '{control_data.get('title', 'Unknown')}': {str(e)}")
                    continue

            self.db.commit()

            if controls_created == 0 and tasks_created == 0:
                return {
                    'success': False,
                    'message': 'No readable tasks were found in the selected sheet. Please choose another tab or verify the sheet structure.',
                    'controls_created': 0,
                    'tasks_created': 0,
                    'errors': errors or ['No readable tasks were found in the selected sheet']
                }

            return {
                'success': True,
                'message': f"Successfully imported {controls_created} controls and {tasks_created} tasks from '{sheet_name or 'active sheet'}'",
                'controls_created': controls_created,
                'tasks_created': tasks_created,
                'errors': errors
            }

        except Exception as e:
            if self.db:
                self.db.rollback()
            return {
                'success': False,
                'message': f'Import failed: {str(e)}',
                'controls_created': 0,
                'tasks_created': 0,
                'errors': [str(e)]
            }

    def _get_or_create_control_template(self, control_data: Dict[str, Any]) -> ControlTemplate:
        """Get existing or create new control template"""
        template = self.db.query(ControlTemplate).filter(
            (ControlTemplate.control_code == control_data['control_code']) |
            (ControlTemplate.title == control_data['title'])
        ).first()

        if template:
            return template

        category = self._extract_category(control_data['control_code'])
        frequency_type = self._determine_frequency_type(
            control_data['frequency_event'],
            control_data['frequency_scheduled']
        )

        template = ControlTemplate(
            title=control_data['title'],
            description=control_data['description'],
            control_code=control_data['control_code'],
            category=category,
            frequency_type=frequency_type,
            event_trigger=control_data['frequency_event'],
            scheduled_frequency=control_data['frequency_scheduled'],
            default_assignees=control_data['assigned_to'],
            evidence_required=True,
            is_active=True
        )

        self.db.add(template)
        self.db.flush()
        return template

    def _create_project_control(self, project_id: int, control_template_id: int,
                                control_data: Dict[str, Any]) -> ProjectControl:
        """Create project control mapping"""
        existing = self.db.query(ProjectControl).filter(
            ProjectControl.project_id == project_id,
            ProjectControl.control_template_id == control_template_id
        ).first()

        if existing:
            if control_data['assigned_to'] and not existing.assigned_to:
                existing.assigned_to = control_data['assigned_to']
            if control_data['evidence_location'] and not existing.evidence_location:
                existing.evidence_location = control_data['evidence_location']
            return existing

        project_control = ProjectControl(
            project_id=project_id,
            control_template_id=control_template_id,
            is_applicable=True,
            assigned_to=control_data['assigned_to'],
            evidence_location=control_data['evidence_location']
        )

        self.db.add(project_control)
        self.db.flush()
        return project_control

    def _create_task_instance(self, project_control_id: int, project_id: int,
                              control_template_id: int, instance_data: Dict[str, Any]) -> TaskInstance:
        """Create or update task instance"""
        existing = self.db.query(TaskInstance).filter(
            TaskInstance.project_control_id == project_control_id,
            TaskInstance.project_id == project_id,
            TaskInstance.control_template_id == control_template_id,
            TaskInstance.instance_label == instance_data['instance_label'],
            TaskInstance.planned_date == instance_data['planned_date']
        ).first()

        if existing:
            metadata_lines = [
                f"Source sheet: {instance_data.get('source_sheet') or ''}",
                f"Source row: {instance_data.get('source_row_number') or ''}",
                f"Task name: {instance_data.get('raw_task_name') or ''}",
                f"Task description: {instance_data.get('raw_task_description') or ''}",
                f"Guidance: {instance_data.get('raw_guidance') or ''}",
                f"Event frequency: {instance_data.get('raw_frequency_event') or ''}",
                f"Scheduled frequency: {instance_data.get('raw_frequency_scheduled') or ''}",
                f"Assigned to: {instance_data.get('raw_assigned_to') or ''}",
                f"Evidence location: {instance_data.get('raw_evidence_location') or ''}",
                f"Planned date raw: {instance_data.get('raw_planned_date') or ''}",
                f"Actual date raw: {instance_data.get('raw_actual_date') or ''}"
            ]
            existing.completion_notes = "\n".join(metadata_lines)
            existing.actual_date = instance_data['actual_date']
            existing.status = instance_data['status']
            existing.evidence_uploaded = bool(instance_data['actual_date'])
            self.db.flush()
            return existing

        metadata_lines = [
            f"Source sheet: {instance_data.get('source_sheet') or ''}",
            f"Source row: {instance_data.get('source_row_number') or ''}",
            f"Task name: {instance_data.get('raw_task_name') or ''}",
            f"Task description: {instance_data.get('raw_task_description') or ''}",
            f"Guidance: {instance_data.get('raw_guidance') or ''}",
            f"Event frequency: {instance_data.get('raw_frequency_event') or ''}",
            f"Scheduled frequency: {instance_data.get('raw_frequency_scheduled') or ''}",
            f"Assigned to: {instance_data.get('raw_assigned_to') or ''}",
            f"Evidence location: {instance_data.get('raw_evidence_location') or ''}",
            f"Planned date raw: {instance_data.get('raw_planned_date') or ''}",
            f"Actual date raw: {instance_data.get('raw_actual_date') or ''}"
        ]

        task_instance = TaskInstance(
            project_control_id=project_control_id,
            project_id=project_id,
            control_template_id=control_template_id,
            instance_label=instance_data['instance_label'],
            planned_date=instance_data['planned_date'],
            actual_date=instance_data['actual_date'],
            status=instance_data['status'],
            completion_notes="\n".join(metadata_lines),
            evidence_uploaded=bool(instance_data['actual_date'])
        )

        self.db.add(task_instance)
        self.db.flush()
        return task_instance

    def _parse_date(self, date_value: Any) -> Optional[datetime]:
        """Parse date from various formats"""
        if date_value in (None, ''):
            return None

        if isinstance(date_value, datetime):
            return date_value

        if hasattr(date_value, 'year') and hasattr(date_value, 'month') and hasattr(date_value, 'day'):
            return datetime(date_value.year, date_value.month, date_value.day)

        if isinstance(date_value, str):
            cleaned_value = self._clean_text(date_value)
            if not cleaned_value:
                return None

            formats = [
                '%d-%b-%y',
                '%d-%m-%Y',
                '%Y-%m-%d',
                '%d/%m/%Y',
                '%d/%m/%y',
                '%d-%b-%Y',
                '%d %b %Y',
                '%d %B %Y',
                '%m/%d/%Y',
            ]

            for fmt in formats:
                try:
                    return datetime.strptime(cleaned_value, fmt)
                except ValueError:
                    continue

        return None

    def _parse_team_members(self, assigned_to: Any) -> List[str]:
        """Parse team members from string"""
        if not assigned_to:
            return []

        if isinstance(assigned_to, list):
            cleaned_members: List[str] = []
            for member in assigned_to:
                cleaned_member = self._clean_text(member)
                if cleaned_member:
                    cleaned_members.append(cleaned_member)
            return cleaned_members

        members = re.split(r'[,;/\n]+', str(assigned_to))
        return [m.strip() for m in members if m.strip()]

    def _generate_instance_label(self, planned_date: Any, frequency: Optional[str]) -> str:
        """Generate instance label like Jan'26 or Q1'26"""
        date = self._parse_date(planned_date)
        if not date:
            return 'Unknown'

        if frequency and 'quarterly' in frequency.lower():
            quarter = (date.month - 1) // 3 + 1
            return f"Q{quarter}'{date.strftime('%y')}"
        if frequency and 'annual' in frequency.lower():
            return date.strftime("%Y")
        return date.strftime("%b'%y")

    def _extract_category(self, control_code: str) -> str:
        """Extract category from control code"""
        if not control_code:
            return 'General'

        match = re.match(r'^([A-Z]+)', control_code)
        if match:
            prefix = match.group(1)
            category_map = {
                'RSK': 'Risk Management',
                'INV': 'Inventory Management',
                'DIP': 'Data Protection',
                'ACC': 'Access Control',
                'SEC': 'Security',
                'AUD': 'Audit'
            }
            return category_map.get(prefix, prefix)

        return 'General'

    def _determine_frequency_type(self, event_driven: Optional[str],
                                  scheduled: Optional[str]) -> str:
        """Determine frequency type"""
        has_event = bool(event_driven and event_driven.strip())
        has_scheduled = bool(scheduled and scheduled.strip())

        if has_event and has_scheduled:
            return 'both'
        if has_event:
            return 'event-driven'
        return 'scheduled'

    def _clean_text(self, value: Any) -> Optional[str]:
        """Normalize text values from Excel"""
        if value is None:
            return None

        cleaned = str(value).replace('\xa0', ' ').strip()
        if not cleaned:
            return None

        if cleaned.lower() in {'none', 'nan', 'null', 'n/a', '-'}:
            return None

        return cleaned

    def _normalize_header(self, value: Any) -> str:
        """Normalize header labels for fuzzy matching"""
        cleaned = self._clean_text(value) or ''
        cleaned = cleaned.lower()
        cleaned = re.sub(r'[^a-z0-9]+', ' ', cleaned)
        return ' '.join(cleaned.split())

    def _stringify_cell(self, value: Any) -> str:
        """Convert Excel cell values to display-friendly strings"""
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        return str(value).strip()

    def _detect_headers(self, sheet: Any) -> tuple[int, Dict[str, int]]:
        """Detect the header row and map columns dynamically"""
        header_aliases = {
            'control_objective': [
                'control objective',
                'control objectives',
                'task',
                'task name',
                'task description',
                'activity',
                'activities'
            ],
            'control_execution_tasks': [
                'control execution tasks',
                'control execution task',
                'execution tasks',
                'details',
                'description'
            ],
            'control_guidance': [
                'control guidance',
                'guidance',
                'reference code',
                'control id',
                'id'
            ],
            'frequency_event_driven': [
                'frequency event driven',
                'event driven',
                'frequency event'
            ],
            'frequency_scheduled': [
                'frequency scheduled',
                'scheduled frequency',
                'frequency'
            ],
            'planned_completion_date': [
                'planned completion date',
                'planned date',
                'target completion date',
                'due date',
                'planned completion',
                'target date'
            ],
            'actual_completion_date': [
                'actual completion date',
                'actual date',
                'completed date'
            ],
            'assigned_to': [
                'assigned to',
                'owner',
                'assignee',
                'assigned'
            ],
            'evidence_storage_location': [
                'evidence storage location',
                'evidence location',
                'evidence storage'
            ]
        }

        best_row_index = 1
        best_mapping: Dict[str, int] = {}
        best_score = -1

        max_rows_to_scan = min(15, getattr(sheet, 'max_row', 15) or 15)
        for row_idx, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=max_rows_to_scan, values_only=True),
            start=1
        ):
            normalized_cells = [self._normalize_header(cell) for cell in row]
            mapping: Dict[str, int] = {}

            for field, aliases in header_aliases.items():
                for index, cell_value in enumerate(normalized_cells):
                    if cell_value in aliases:
                        mapping[field] = index
                        break

            if len(mapping) > best_score:
                best_score = len(mapping)
                best_mapping = mapping
                best_row_index = row_idx

        if best_score < 2:
            raise ValueError(
                "Could not detect the required Excel columns in the selected tab. "
                "Please choose another sheet that contains task names and due dates."
            )

        if 'planned_completion_date' not in best_mapping:
            raise ValueError("The selected sheet must contain a due date or planned completion date column.")

        if 'control_objective' not in best_mapping:
            best_mapping['control_objective'] = best_mapping.get('control_execution_tasks', 0)

        return best_row_index, best_mapping

    def _get_cell_value(self, row: Any, headers: Dict[str, int], field: str) -> Any:
        """Safely read a cell value using detected headers"""
        column_index = headers.get(field)
        if column_index is None or column_index >= len(row):
            return None
        return row[column_index]


# Made with Bob
